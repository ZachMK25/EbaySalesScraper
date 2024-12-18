from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import NoSuchElementException, ElementNotInteractableException, TimeoutException
import datetime
from bs4 import BeautifulSoup
import re
import openpyxl
from datetime import datetime

def convert_to_year_month_day(month_day: str) -> str:

    current_date = datetime.now()
    current_year = current_date.year

    input_date = datetime.strptime(
        month_day, "%b %d"
    )  # "%b" = month abbreviation, "%d" = day

    input_month = input_date.month
    input_day = input_date.day

    if input_month > current_date.month or (
        input_month == current_date.month and input_day > current_date.day
    ):
        year = current_year - 1
    else:
        year = current_year

    result_date = datetime(year, input_month, input_day)
    return result_date.strftime("%Y-%m-%d")


def gross_date_conversion(date):  # expects date in "MMM DD" format (i.e. "Dec 14"), including space in between
    # outputs in YYYY-MM-DD of the most recent date
    return convert_to_year_month_day(" ".join(date.text.strip().replace("\n", "").split(" ")[-2:]))


def run():

    columns = {"title": "B", "date": "G", "price": "H", "shipping": "I"}

    shipping_regex = r"buyer paid \$([0-9]+\.[0-9]{2})"

    login = True

    if login:

        driver = webdriver.Chrome()

        driver.get("https://signin.ebay.com")

        print("Please log in manually.")
        try:
            # signed in
            # gh-identity__greeting

            # signed out
            # gh-identity-signed-out-recognized
            
            errors = [NoSuchElementException]
            # Wait up to 5 minutes for the CAPTCHA to be solved
            wait = WebDriverWait(driver, 300, poll_frequency=0.2, ignored_exceptions=errors).until(EC.visibility_of_element_located((By.XPATH, "//span[contains(@class, 'gh-identity__greeting')]")))
            
            print("Element found:", wait.text)

            print("CAPTCHA solved! Proceeding...")

            driver.get(
                "https://www.ebay.com/mys/sold/rf/sort=MOST_RECENTLY_SOLD&filter=ALL&limit=200&period=LAST_90_DAYS"
            )

            driver.implicitly_wait(10)  # seconds

            sold_items = driver.page_source

        except TimeoutException:
            # print(e)
            print("CAPTCHA and Login took too long or failed.")
            driver.quit()
            exit(0)

    else:
        input_file = open("example.html", "r")

        html = input_file.read()

        input_file.close()

        sold_items = html


    soup = BeautifulSoup(sold_items, "lxml")

    # sold-item--content div --> listing

    # "item-title" --> descriptions
    # "item__sold-date" --> date sold
    # "item__shipping-price" --> shipping price [+ Shipping (buyer paid $X)] or [+ Shipping]
    # "item__price" --> sale price
    # "item__buyer-name" --> buyer name

    listings = soup.find_all("div", class_="sold-item--content")

    # cleaned_listings = []

    f = open("output.txt", "w")

    excel_file_name = "INVENTORY & SALES TRACKING.xlsx"

    wb = openpyxl.load_workbook(excel_file_name)
    ws = wb.active

    row_counter = 6  # arbitrary starting point of spreadsheet to make it look cleaner

    for listing in listings:
        # print (listing)
        try:
            title = listing.find("h3", class_="item-title")
            date = listing.find("div", class_="item__sold-date")
            price = listing.find("span", class_="item__price")
            shipping = listing.find("span", class_="item__shipping-price")
            # buyer_name = listing.find("div", class_="item__buyer-name")

            if shipping:
                match = re.search(shipping_regex, shipping.text.strip())

                if match:
                    shipping = "$" + match.group(1)
                else:
                    shipping = "$0"

            listing_dict = {
                "title": " ".join(title.text.strip().split()) if title else "NO TITLE",
                "date": gross_date_conversion(date) if date else "NO DATE",
                "price": price.text.strip().replace("\n", "") if price else "NO PRICE",
                "shipping": (
                    shipping.strip().replace("\n", "") if shipping else "SHIPPING NOT FOUND"
                ),
                # "buyer_name": buyer_name.text.strip() if buyer_name else "NO BUYER"
            }

            ws[columns["title"] + str(row_counter)] = listing_dict["title"]

            ws[columns["date"] + str(row_counter)] = listing_dict["date"]

            ws[columns["price"] + str(row_counter)] = listing_dict["price"]
            ws[columns["price"] + str(row_counter)].number_format = (
                "$#,##0.00"  # set the format to USD
            )

            ws[columns["shipping"] + str(row_counter)] = listing_dict["shipping"]
            ws[columns["shipping"] + str(row_counter)].number_format = (
                "$#,##0.00"  # set the format to USD
            )

            row_counter += 1
            f.write(str(listing_dict) + "\n\n")

        except Exception as e:
            print(f"Error parsing listing: {e}")

    f.close()

    wb.save(".xlsx")

    if login:
        driver.quit()